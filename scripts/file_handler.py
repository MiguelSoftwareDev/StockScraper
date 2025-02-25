import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from logger_config import logger


def salvar_parquet(dados, nome_arquivo):
    """
    Salva os dados em um arquivo Parquet.
    Saves data to a Parquet file.

    Args:
        dados (list): Lista de dicionários com os dados.
                      List of dictionaries with data.
        nome_arquivo (str): Nome do arquivo Parquet.
                            Name of the Parquet file.

    Returns:
        str: Caminho completo do arquivo salvo.
             Full path of the saved file.
    """
    try:
        caminho_documentos = Path.home() / "Documents" / "BolsaValores" / "exports"
        caminho_documentos.mkdir(parents=True, exist_ok=True)
        caminho_completo = caminho_documentos / nome_arquivo
        df = pd.DataFrame(dados)
        df.to_parquet(caminho_completo, compression="snappy")
        logger.info(f"Dados salvos em Parquet: {caminho_completo} | Data saved to Parquet: {caminho_completo}")
        return caminho_completo
    except Exception as e:
        logger.error(f"Erro ao salvar dados em Parquet: {e} | Error saving data to Parquet: {e}")
        return None

def salvar_excel_formatado(dados_acoes, dados_moedas, nome_arquivo):
    """
    Salva os DataFrames em um arquivo Excel com formatação profissional.
    Saves the DataFrames to an Excel file with professional formatting.

    Args:
        dados_acoes (list): Lista de dicionários com os dados das ações.
                            List of dictionaries with stock data.
        dados_moedas (list): Lista de dicionários com os dados das moedas.
                             List of dictionaries with currency data.
        nome_arquivo (str): Nome do arquivo Excel.
                            Name of the Excel file.

    Returns:
        str: Caminho completo do arquivo salvo.
             Full path of the saved file.
    """
    try:
        caminho_documentos = Path.home() / "Documents" / "BolsaValores" / "exports"
        caminho_documentos.mkdir(parents=True, exist_ok=True)
        caminho_completo = caminho_documentos / nome_arquivo

        # Converter listas de dicionários para DataFrames
        df_acoes = pd.DataFrame(dados_acoes)
        df_moedas = pd.DataFrame(dados_moedas)

        # Salvar em Excel
        with pd.ExcelWriter(caminho_completo, engine="openpyxl") as writer:
            df_acoes.to_excel(writer, index=False, sheet_name="Ações")
            df_moedas.to_excel(writer, index=False, sheet_name="Moedas")

            # Formatar a planilha de ações | Format the stock sheet
            workbook = writer.book
            worksheet_acoes = writer.sheets["Ações"]
            
            # Ajustar largura das colunas | Adjust column width
            for col in worksheet_acoes.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                worksheet_acoes.column_dimensions[col[0].column_letter].width = max_length + 2
            
            # Centralizar texto | Center text
            for row in worksheet_acoes.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            
            # Negrito no cabeçalho | Bold header
            for cell in worksheet_acoes[1]:
                cell.font = Font(bold=True)
            
            # Adicionar bordas | Add borders
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            for row in worksheet_acoes.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Formatação condicional para tendências em "Ações" | Conditional formatting for trends in "Stocks"
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            worksheet_acoes.conditional_formatting.add(
                f"F2:F{len(df_acoes) + 1}",  # Coluna "TENDÊNCIA" | Column "TREND"
                CellIsRule(operator="equal", formula=['"Subindo | Up"'], fill=green_fill)
            )
            worksheet_acoes.conditional_formatting.add(
                f"F2:F{len(df_acoes) + 1}",  # Coluna "TENDÊNCIA" | Column "TREND"
                CellIsRule(operator="equal", formula=['"Caindo | Down"'], fill=red_fill)
            )

            # Formatar a planilha de moedas | Format the currency sheet
            worksheet_moedas = writer.sheets["Moedas"]
            
            # Ajustar largura das colunas | Adjust column width
            for col in worksheet_moedas.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                worksheet_moedas.column_dimensions[col[0].column_letter].width = max_length + 2
            
            # Centralizar texto | Center text
            for row in worksheet_moedas.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            
            # Negrito no cabeçalho | Bold header
            for cell in worksheet_moedas[1]:
                cell.font = Font(bold=True)
            
            # Adicionar bordas | Add borders
            for row in worksheet_moedas.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Formatação condicional para moedas | Conditional formatting for currencies
            worksheet_moedas.conditional_formatting.add(
                f"E2:E{len(df_moedas) + 1}",  # Coluna "TENDÊNCIA" | Column "TREND"
                CellIsRule(operator="equal", formula=['"Subindo | Up"'], fill=green_fill)
            )
            worksheet_moedas.conditional_formatting.add(
                f"E2:E{len(df_moedas) + 1}",  # Coluna "TENDÊNCIA" | Column "TREND"
                CellIsRule(operator="equal", formula=['"Caindo | Down"'], fill=red_fill)
            )

        logger.info(f"Dados exportados para {caminho_completo} com sucesso! | Data exported to {caminho_completo} successfully!")
        return caminho_completo
    except Exception as e:
        logger.error(f"Erro ao exportar dados para Excel: {e} | Error exporting data to Excel: {e}")
        return None

def abrir_excel(caminho):
    """
    Abre o arquivo Excel automaticamente com base no sistema operacional.
    Opens the Excel file automatically based on the operating system.

    Args:
        caminho (str): Caminho do arquivo Excel.
                       Path to the Excel file.
    """
    try:
        if sys.platform.startswith("win"):  # Windows
            os.startfile(caminho)
        elif sys.platform.startswith("darwin"):  # macOS
            os.system(f'open "{caminho}"')
        else:  # Linux e outros | Linux and others
            os.system(f'xdg-open "{caminho}"')
        logger.info(f"Arquivo Excel aberto automaticamente: {caminho} | Excel file opened automatically: {caminho}")
    except Exception as e:
        logger.error(f"Erro ao abrir o arquivo Excel: {e} | Error opening Excel file: {e}")